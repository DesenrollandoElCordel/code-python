from openpyxl import load_workbook
from lxml import etree
import xml.etree.ElementTree as eT
import os

# Creation of a list with <term>
def add_list_keywords(keyword):
    # Path to the XML taxonomy
    tree = eT.parse('../Encodage/TEI_tests/taxonomy.xml')
    root = tree.getroot()
    ns = {'tei': 'http://www.tei-c.org/ns/1.0'}
    list_keywords = etree.SubElement(keywords, "term")

    # For each keyword in the tabler, we check if it exists in the taxonomy and put it in <term> element
    for x in root.findall('.//tei:category', ns):
        att = x.get('{http://www.w3.org/XML/1998/namespace}id')
        catdesc = x[0].text
        if att == keyword:
            list_keywords.text = catdesc


# Creation of a list of keywords with <catRef>
def add_keywords(keyword, category, element):
    if keyword is not None:
        if "," in keyword:  # For categories with several keywords
            keyword_split = keyword.split(",")
            for k in keyword_split:
                catref = etree.SubElement(element, "catRef", scheme="#" + category, target="#" + k)
                add_list_keywords(k)
        else:
            catref = etree.SubElement(element, "catRef", scheme="#" + category, target="#" + keyword)
            add_list_keywords(keyword)


ns = {'tei': 'http://www.tei-c.org/ns/1.0'}  #  Namespace declaration

wb = load_workbook(filename='../varios_woodcuts_update.xlsx')  # Workbook loading
index = wb['metadata']  # Name of the worksheet
# print(index)

# We get the columns name
list_name_columns = []
for row in index.iter_rows(max_row=1, max_col=24, values_only=True):
    list_name_columns.append(row)
print(list_name_columns)

# We put the data of the rows in a list
title_engravings = []
for row in index.iter_rows(min_row=131, max_row=954, max_col=24, values_only=True):
    row_list = list(row)
    title_engravings.append(row_list)
# print(title_engravings)

# Creation of the XML-TEI files (One per row)
for i in range(len(title_engravings)):
    root = etree.Element("TEI", xmlns="http://www.tei-c.org/ns/1.0")
    root.set("{http://www.w3.org/XML/1998/namespace}id", title_engravings[i][0])
    teiHeader = etree.SubElement(root, "teiHeader")

    fileDesc = etree.SubElement(teiHeader, "fileDesc")
    titleStmt = etree.SubElement(fileDesc, "titleStmt")
    title = etree.SubElement(titleStmt, "title")
    title.text = title_engravings[i][0]

    respStmt = etree.SubElement(titleStmt, "respStmt")
    name = etree.SubElement(respStmt, "name")
    name.set("{http://www.w3.org/XML/1998/namespace}id", "EL")
    name.text = "Elina Leblanc"
    resp = etree.SubElement(respStmt, "resp")
    resp.text = "Codificación TEI"

    pubStmt = etree.SubElement(fileDesc, "publicationStmt")
    pubStmt_child1 = etree.SubElement(pubStmt, "authority")
    pubStmt_child1.text = "Bibliothèque Universitaire de Genève (BUNIGE)"
    pubStmt_child2 = etree.SubElement(pubStmt, "availability", status="restricted")
    licence = etree.SubElement(pubStmt_child2, "licence", n="cc by nc sa",
                               target="https://creativecommons.org/licenses/by-nc-sa/4.0/")

    sourceDesc = etree.SubElement(fileDesc, "sourceDesc")
    bibl = etree.SubElement(sourceDesc, "bibl")

    title_bibl = etree.SubElement(bibl, "title")
    title_bibl.text = title_engravings[i][0]
    author = etree.SubElement(bibl, "author")
    author.text = title_engravings[i][5]

    publisher = etree.SubElement(bibl, "publisher")
    publisher.text = title_engravings[i][7]

    pubPlace = etree.SubElement(bibl, "pubPlace", ref="xxxx")
    pubPlace.text = title_engravings[i][8]

    date = etree.SubElement(bibl, "date")
    date_text = str(title_engravings[i][6])
    date.text = date_text
    if title_engravings[i][6] == "[s.a.]":
        date.set("when", "1800")
        date.set("cert", "low")
    elif "[?]" in date_text:
        date_medium = title_engravings[i][6][:4]
        date.set("when", date_medium)
        date.set("cert", "medium")
    else:
        date.set("when", date_text)
        date.set("cert", "high")

    ident = etree.SubElement(bibl, "ident", type="DOI")

    figure = etree.SubElement(bibl, "figure")

    coords = title_engravings[i][3].split(',')
    width = int(coords[2]) - int(coords[0])
    height = int(coords[3]) - int(coords[1])
    new_coords = coords[0] + "," + coords[1] + "," + str(width) + "," + str(height)
    # print(new_coords)

    graphic = etree.SubElement(figure, "graphic", url="varios/" + title_engravings[i][2] + ".jpg/" + new_coords + "/full/0/default.jpg",
                               source=title_engravings[i][2] + ".jpg",
                               n=new_coords)

    figDesc = etree.SubElement(figure, "figDesc")
    locus1 = etree.SubElement(figDesc, "locus", target=title_engravings[i][1]+".xml")
    locus1.text = title_engravings[i][4]

    note1 = etree.SubElement(bibl, "note", type="similar_ejemplar")
    if title_engravings[i][23] is not None:
        list_sameAs = etree.SubElement(note1, "list")
        if "," in title_engravings[i][23]:
            sameAs_split = title_engravings[i][23].split(",")
            list_sameAs.set("n", str(len(sameAs_split)))
            for s in sameAs_split:
                item = etree.SubElement(list_sameAs, "item")
                title_sameAs = etree.SubElement(item, "title", corresp=s + ".xml")
                title_sameAs.text = s
                bibl2 = etree.SubElement(item, "bibl")
                title2 = etree.SubElement(bibl2, "title")
                date2 = etree.SubElement(bibl2, "date")
                publisher2 = etree.SubElement(bibl2, "publisher")
        else:
            list_sameAs.set("n", "1")
            item = etree.SubElement(list_sameAs, "item")
            title_sameAs = etree.SubElement(item, "title", corresp=title_engravings[i][23] + '.xml')
            title_sameAs.text = title_engravings[i][23]
            bibl2 = etree.SubElement(item, "bibl")
            title2 = etree.SubElement(bibl2, "title")
            date2 = etree.SubElement(bibl2, "date")
            publisher2 = etree.SubElement(bibl2, "publisher")

    note2 = etree.SubElement(bibl, "note", type="origen")
    note2.text = "Grabado extraído del pliego " + title_engravings[i][1]

    encodingDesc = etree.SubElement(teiHeader, "encodingDesc")
    projectDesc = etree.SubElement(encodingDesc, "projectDesc")
    paragraph = etree.SubElement(projectDesc, "p")
    paragraph.text = "Este archivo fue creado en el marco del proyecto Desenrollando el Cordel/Démêler le cordel/Untangling the cordel, dirigido por la profesora Constance Carta de la Universidad de Ginebra, con el apoyo de la Fundación filantrópica Famille Sandoz-Monique de Meuron."

    profileDesc = etree.SubElement(teiHeader, "profileDesc")
    textClass = etree.SubElement(profileDesc, 'textClass')
    keywords = etree.SubElement(textClass, 'keywords')

    add_keywords(title_engravings[i][9], list_name_columns[0][9], textClass)  # Category 'personaje_masculino'
    add_keywords(title_engravings[i][10], list_name_columns[0][10], textClass)  # Category 'personaje_femenino'
    add_keywords(title_engravings[i][11], list_name_columns[0][11], textClass)  # Category 'ninos'
    add_keywords(title_engravings[i][12], list_name_columns[0][12], textClass)  # Category 'grupos de personajes'
    add_keywords(title_engravings[i][13], list_name_columns[0][13], textClass)  # Category 'accion'
    add_keywords(title_engravings[i][14], list_name_columns[0][14], textClass)  # Category 'muerte'
    add_keywords(title_engravings[i][15], list_name_columns[0][15], textClass)  # Category 'religion'
    add_keywords(title_engravings[i][16], list_name_columns[0][16], textClass)  # Category 'animales'
    add_keywords(title_engravings[i][17], list_name_columns[0][17], textClass)  # Category 'accesorios'
    add_keywords(title_engravings[i][18], list_name_columns[0][18], textClass)  # Category 'construidos'
    add_keywords(title_engravings[i][19], list_name_columns[0][19], textClass)  # Category 'naturales'
    add_keywords(title_engravings[i][20], list_name_columns[0][20], textClass)  # Category 'transporte'
    add_keywords(title_engravings[i][21], list_name_columns[0][21], textClass)  # Category 'escudo'
    add_keywords(title_engravings[i][22], list_name_columns[0][22], textClass)  # Category 'ornamentos'

    revisionDesc = etree.SubElement(teiHeader, "revisionDesc")
    change = etree.SubElement(revisionDesc, "change", who="#EL", when="2021-11-22")
    change.text = "Creación del documento"

    facsimile = etree.SubElement(root, "facsimile")
    thumbnail = etree.SubElement(facsimile, "graphic", url="varios/" + title_engravings[i][2] + ".jpg/" + new_coords + "/200,/0/default.jpg")

    text = etree.SubElement(root, "text")
    body = etree.SubElement(text, "body")
    paragraph2 = etree.SubElement(body, "p")

    tree = etree.ElementTree(root)  # Creation of the XML tree

    path_tei = '../test_gravures'  # New folder to save the files
    if not os.path.isdir(path_tei):
        os.mkdir(path_tei)

    filename = title_engravings[i][0] + '.xml'  # New name of the files

    filename_path = os.path.join(path_tei, filename)  # Creation of the path
    # print(filename_path)

    if not os.path.isfile(filename_path):
        tree.write(filename_path, xml_declaration=True, encoding='UTF-8', pretty_print=True)

    print(title_engravings[i][0] + " => done")
    # print(etree.tostring(root, xml_declaration=True, encoding='UTF-8', pretty_print=True))
