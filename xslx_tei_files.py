from openpyxl import load_workbook
from lxml import etree
import xml.etree.ElementTree as eT
import os


# Fonction pour la gestion des mots-clés
def add_list_keywords(keyword):
    # Chemin vers la taxonomie au format .xml
    tree = eT.parse('/Users/elinaleblanc/Documents/Postdoctorat/Encodage/TEI_tests/Index_gravures/taxonomy_grabados.xml')
    root = tree.getroot()
    ns = {'tei': 'http://www.tei-c.org/ns/1.0'}
    list_keywords = etree.SubElement(keywords, "term")  # Création d'un <term> pour chaque mot-clé

    # Pour chaque mot-clé du tableau, on vérifie s'il existe dans la taxonomie et on l'ajoute avec un <term>
    for x in root.findall('.//tei:category', ns):
        att = x.get('{http://www.w3.org/XML/1998/namespace}id')
        catdesc = x[0].text
        if att == keyword:
            list_keywords.text = catdesc


def add_keywords(keyword, category, element):
    if keyword is not None:
        if "," in keyword:  # Gestion des catégories avec plusieurs mots-clés
            keyword_split = keyword.split(",")
            for k in keyword_split:
                catref = etree.SubElement(element, "catRef", scheme="#" + category, target="#" + k)
                add_list_keywords(k)
        else:
            catref = etree.SubElement(element, "catRef", scheme="#" + category, target="#" + keyword)
            add_list_keywords(keyword)


tree_index_imp = eT.parse("/Users/elinaleblanc/Documents/Postdoctorat/Encodage/impresor.xml")
root_index_imp = tree_index_imp.getroot()

tree_index_lugar = eT.parse("/Users/elinaleblanc/Documents/Postdoctorat/Encodage/lugar_impresion.xml")
root_index_lugar = tree_index_lugar.getroot()

ns = {'tei': 'http://www.tei-c.org/ns/1.0'}

wb = load_workbook(filename='/Users/elinaleblanc/Documents/Postdoctorat/Index_Pliegos/Index_Grabados_Moreno.xlsx')
index = wb['Feuil1']
# print(index)

# Récupération des noms de colonnes
list_name_columns = []
for row in index.iter_rows(max_row=1, max_col=25, values_only=True):
    list_name_columns.append(row)
# print(list_name_columns)

# Récupération des descriptions
title_engravings = []
for row in index.iter_rows(min_row=2, max_row=619, max_col=26, values_only=True):
    row_list = list(row)
    title_engravings.append(row_list)
# print(title_engravings)

# Création des fichiers .xml
for i in range(len(title_engravings)):
    root = etree.Element("TEI", xmlns="http://www.tei-c.org/ns/1.0")
    root.set("{http://www.w3.org/XML/1998/namespace}id", title_engravings[i][0])
    teiHeader = etree.SubElement(root, "teiHeader")

    fileDesc = etree.SubElement(teiHeader, "fileDesc")
    titleStmt = etree.SubElement(fileDesc, "titleStmt")
    title = etree.SubElement(titleStmt, "title")
    title.text = title_engravings[i][0]

    respStmt = etree.SubElement(titleStmt, "respStmt")
    name = etree.SubElement(respStmt, "name")
    name.set("{http://www.w3.org/XML/1998/namespace}id", "EL")
    name.text = "Elina Leblanc"
    resp = etree.SubElement(respStmt, "resp")
    resp.text = "Codificación TEI"

    pubStmt = etree.SubElement(fileDesc, "publicationStmt")
    pubStmt_child1 = etree.SubElement(pubStmt, "authority")
    pubStmt_child1.text = "Bibliothèque Universitaire de Genève (BUNIGE)"
    pubStmt_child2 = etree.SubElement(pubStmt, "availability", status="restricted")
    licence = etree.SubElement(pubStmt_child2, "licence", n="cc by nc sa",
                               target="https://creativecommons.org/licenses/by-nc-sa/4.0/")

    sourceDesc = etree.SubElement(fileDesc, "sourceDesc")
    bibl = etree.SubElement(sourceDesc, "bibl")

    title_bibl = etree.SubElement(bibl, "title")
    title_bibl.text = title_engravings[i][0]
    author = etree.SubElement(bibl, "author")

    publisher = etree.SubElement(bibl, "publisher")
    for impresor in root_index_imp.findall(".//tei:person", ns):
        attribute_impresor = impresor.get('{http://www.w3.org/XML/1998/namespace}id')
        name_impresor = impresor[0].text
        # print(name_impresor)
        if attribute_impresor == title_engravings[i][6]:
            publisher.text = name_impresor
            publisher.set("corresp", "impresor.xml#" + title_engravings[i][6])

    pubPlace = etree.SubElement(bibl, "pubPlace", ref="https://www.geonames.org/2520118/carmona.html")
    for lugar in root_index_lugar.findall(".//tei:place", ns):
        attribute_lugar = lugar.get('{http://www.w3.org/XML/1998/namespace}id')
        name_lugar = lugar[0].text
        if attribute_lugar == title_engravings[i][7]:
            pubPlace.text = name_lugar
            pubPlace.set("corresp", "lugar.xml#" + title_engravings[i][7])

    date = etree.SubElement(bibl, "date")
    date_text = str(title_engravings[i][5])
    date.text = date_text
    if title_engravings[i][5] == "[s.a.]":
        date.set("when", "1800")
        date.set("cert", "low")
    elif "[?]" in date_text:
        date_medium = title_engravings[i][5][:4]
        date.set("when", date_medium)
        date.set("cert", "medium")
    else:
        date.set("when", date_text)
        date.set("cert", "high")

    ident = etree.SubElement(bibl, "ident", type="DOI")

    figure = etree.SubElement(bibl, "figure")

    graphic = etree.SubElement(figure, "graphic", url="",
                               source=title_engravings[i][2] + ".jpg",
                               n=title_engravings[i][3][1:-1])

    figDesc = etree.SubElement(figure, "figDesc")
    locus1 = etree.SubElement(figDesc, "locus", target=title_engravings[i][1]+".xml")
    locus1.text = title_engravings[i][4]

    note1 = etree.SubElement(bibl, "note", type="similar_ejemplar")
    if title_engravings[i][25] is not None:
        list_sameAs = etree.SubElement(note1, "list")
        if "," in title_engravings[i][25]:
            sameAs_split = title_engravings[i][25].split(",")
            list_sameAs.set("n", str(len(sameAs_split)))
            for s in sameAs_split:
                item = etree.SubElement(list_sameAs, "item")
                title_sameAs = etree.SubElement(item, "title", corresp=s + ".xml")
                title_sameAs.text = s
                locus2 = etree.SubElement(item, "locus")
                date2 = etree.SubElement(item, "date")
        else:
            list_sameAs.set("n", "1")
            item = etree.SubElement(list_sameAs, "item")
            title_sameAs = etree.SubElement(item, "title", corresp=title_engravings[i][25] + '.xml')
            title_sameAs.text = title_engravings[i][25]
            locus2 = etree.SubElement(item, "locus")
            date2 = etree.SubElement(item, "date")

    note2 = etree.SubElement(bibl, "note", type="origen")
    note2.text = "Grabado extraído del pliego " + title_engravings[i][1]

    encodingDesc = etree.SubElement(teiHeader, "encodingDesc")
    projectDesc = etree.SubElement(encodingDesc, "projectDesc")
    paragraph = etree.SubElement(projectDesc, "p")
    paragraph.text = "Este archivo fue creado en el marco del proyecto Desenrollando el Cordel/Démêler le cordel/Untangling the cordel, dirigido por la profesora Constance Carta de la Universidad de Ginebra, con el apoyo de la Fundación filantrópica Famille Sandoz-Monique de Meuron."

    profileDesc = etree.SubElement(teiHeader, "profileDesc")
    textClass = etree.SubElement(profileDesc, 'textClass')
    keywords = etree.SubElement(textClass, 'keywords')

    add_keywords(title_engravings[i][8], list_name_columns[0][8], textClass)  # Catégorie 'personaje_masculino'
    add_keywords(title_engravings[i][9], list_name_columns[0][9], textClass)  # Catégorie 'personaje_femenino'
    add_keywords(title_engravings[i][10], list_name_columns[0][10], textClass)  # Catégorie 'grupos_personajes'
    add_keywords(title_engravings[i][11], list_name_columns[0][11], textClass)  # Catégorie 'actitud'
    add_keywords(title_engravings[i][12], list_name_columns[0][12], textClass)  # Catégorie 'muerte'
    add_keywords(title_engravings[i][13], list_name_columns[0][13], textClass)  # Catégorie 'religion'
    add_keywords(title_engravings[i][14], list_name_columns[0][14], textClass)  # Catégorie 'monstruo'
    add_keywords(title_engravings[i][15], list_name_columns[0][15], textClass)  # Catégorie 'animales'
    add_keywords(title_engravings[i][16], list_name_columns[0][16], textClass)  # Catégorie 'atuendo'
    add_keywords(title_engravings[i][17], list_name_columns[0][17], textClass)  # Catégorie 'instrumento_musical'
    add_keywords(title_engravings[i][18], list_name_columns[0][18], textClass)  # Catégorie 'arma_de_fuego'
    add_keywords(title_engravings[i][19], list_name_columns[0][19], textClass)  # Catégorie 'arma_blanca'
    add_keywords(title_engravings[i][20], list_name_columns[0][20], textClass)  # Catégorie 'accesorios_varios'
    add_keywords(title_engravings[i][21], list_name_columns[0][21], textClass)  # Catégorie 'espacio_construido'
    add_keywords(title_engravings[i][22], list_name_columns[0][22], textClass)  # Catégorie 'ambiente_natural'
    add_keywords(title_engravings[i][23], list_name_columns[0][23], textClass)  # Catégorie 'ambiente_maritimo'
    add_keywords(title_engravings[i][24], list_name_columns[0][24], textClass)  # Catégorie 'elementos_decorativos'

    revisionDesc = etree.SubElement(teiHeader, "revisionDesc")
    change = etree.SubElement(revisionDesc, "change", who="#EL", when="2021-11-22")
    change.text = "Creación del documento"

    facsimile = etree.SubElement(root, "facsimile")
    thumbnail = etree.SubElement(facsimile, "graphic", url=title_engravings[i][0] + "_thumbnail.jpg")

    text = etree.SubElement(root, "text")
    body = etree.SubElement(text, "body")
    paragraph2 = etree.SubElement(body, "p")

    tree = etree.ElementTree(root)

    path_tei = '../TEI_Gravures/'
    if not os.path.isdir(path_tei):
        os.mkdir(path_tei)

    filename = title_engravings[i][0] + '.xml'

    filename_path = os.path.join(path_tei, filename)
    # print(filename_path)

    if not os.path.isfile(filename_path):
        tree.write(filename_path, xml_declaration=True, encoding='UTF-8', pretty_print=True)

    print(title_engravings[i][0] + " => done")
    # print(etree.tostring(root, xml_declaration=True, encoding='UTF-8', pretty_print=True))
