import os
import argparse
import cv2
import xml.etree.ElementTree as et
import re
import numpy as np
from openpyxl import Workbook

# Arguments declaration
parser = argparse.ArgumentParser()
parser.add_argument('--xml', type=str, help="Path to the XML files to load")
args = parser.parse_args()

# Creation of the XLSX file
wb = Workbook()
wb_filename = ""  # Name of the Workbook

# Creation of the worksheet
ws2 = wb.active
ws2.title = "metadata"

list_woodcuts = []

for subdir, dirs, files in os.walk(args.xml):
    for file in files:
        xml_path = os.path.join(subdir, file)  # Path to the PAGE-XML files
        label, ext = os.path.splitext(file)  # Splitting of the filename

        ns = {'page': 'http://schema.primaresearch.org/PAGE/gts/pagecontent/2013-07-15'}  # Namespace declaration

        if file.endswith('.xml'):
            document = et.parse(xml_path)  # We parse the XML file
            root = document.getroot()  # We get the root

            list_coord_str = []  # List with the raw coordinates (str)
            list_coord_int = []  # List with the coordinates transformed in int

            woodcutCounter = 0  # Initialisation of the number of woodcuts per file after each iteration

            # Extraction of the woodcuts coordinates
            for coord in root.findall(".//page:GraphicRegion/page:Coords", ns):
                point = coord.get('points')
                point = point.replace(',', ' ')
                point_regex = re.split("\s", point)
                list_coord_str.append(point_regex)

            # We convert the coordinated in integer
            for i in range(len(list_coord_str)):
                list_coord_int.append([])  # Nested list for the coordinates
                for j in list_coord_str[i]:
                    number = int(j)
                    list_coord_int[i].append(number)

            # We convert the coordinates in the format (X start, Y start, X end, Y end)
            for k in range(len(list_coord_int)):
                points = np.array([[[list_coord_int[k][0], list_coord_int[k][1]],
                                    [list_coord_int[k][2], list_coord_int[k][3]],
                                    [list_coord_int[k][4], list_coord_int[k][5]],
                                    [list_coord_int[k][6], list_coord_int[k][7]]]])
                rect = cv2.boundingRect(points)
                # print(label, rect)

                # We get information about the file
                for page in root.findall('.//page:Page', ns):
                    documentName = page.get('imageFilename')
                    pliegoName = documentName[:10]  # Document identifier
                    pageName = re.split("\.", documentName)  # Page number
                    woodcutCounter = woodcutCounter + 1  # Incrementation of the number of woodcuts
                    woodcutName = 'grabado_m' + documentName[6:11] + str(woodcutCounter)  # Woodcut identifier
                    # print(woodcutName)

                    # All the information are put in a nested list
                    list_woodcuts.append([woodcutName, pliegoName, pageName[0], str(rect)])

                    # From the information in list_woodcuts, we create the Excel file
                    row = 1  # Initialisation of the number of row
                    for z in list_woodcuts:
                        ws2.cell(column=1, row=row, value=z[0])
                        ws2.cell(column=2, row=row, value=z[1])
                        ws2.cell(column=3, row=row, value=z[2])
                        ws2.cell(column=4, row=row, value=z[3])
                        row += 1  # Incrementation of the row

print(list_woodcuts)
# wb.save(filename=wb_filename)
