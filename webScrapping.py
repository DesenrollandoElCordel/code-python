from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import requests
from openpyxl import Workbook
import json

# Creation of the XLSX file
wb = Workbook()
wb_filename = "../cudl_data.xlsx"

# Creation of the WorkSheet
ws = wb.active

# Base URL of the website
url_base = 'https://cudl.lib.cam.ac.uk/collections/spanishchapbooks/'

item_url_list = []  # List with documents URLs
item_data = []  # List with data about documents (Manifest, title, country, city, date, printer)
manifests_url = []  # List with the manifests URLs

# We browse the pages with the list of documents (Range = number of pages to browse)
for url in range(7, 8):
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))  # Installation of the WebDriver
    driver.get(url_base + str(url))  # URL to be scrapped (Concatenation of Base URL and the nb of pages)
    time.sleep(5)  # Number of seconds before the Web page is closed
    html = driver.page_source  # We get the HTML
    soup = BeautifulSoup(html, "html.parser")  # We parse the page with BeautifulSoup

    # Extraction of document URLs
    item_page = soup.find_all('div', {'class': 'collections_carousel_text'})
    for i in item_page:
        item_page_link = i.find('a')
        item_url_list.append('https://cudl.lib.cam.ac.uk' + item_page_link.get('href'))

# Extraction of the document manifest from the document page
for item_url in item_url_list:
    req = requests.get(item_url)   # GET request on each URL in item_url_list
    soup2 = BeautifulSoup(req.text, 'html.parser')  # We parse the HTML
    iiif_button = soup2.find('div', {'id': 'iiifOption'})  # Button with the manifest URL
    iiif_div = iiif_button.find('div', {'class': 'button'})
    iiif_url = iiif_button.find('a', {'class': 'btn'}).get('href')
    manifest = iiif_url.split("?")  # We split the URL to only have the portion corresponding to the manifest
    manifests_url.append(manifest[1][9:])  # We put the manifest in a list

for m in manifests_url:
    jsonFile = json.loads(requests.get(m).text)  # We parse each manifest contained in the list manifests_url

    # Declaration of the variables
    title = ""
    country = ""
    city = ""
    date = ""
    printer = ""

    # We search for specific information in the Metadata section of each manifest
    # We parse the dictionary to only get the values = dict_values(['label', 'value'])
    for i in jsonFile["metadata"]:
        # Conversion of the dict in list
        # The 1st element of the list is the name of the metadata (Ex: Date of Publication, Title, etc.)
        # The 2nd element of the list is the value of this metadata
        if "Place of Publication" in i.values():
            places = list(i.values())
            # We get the second element of the list (Ex: 'Spain; S.l.')
            # We split it to have the country one side and the city on the other side
            place = places[1].split(";")
            city = place[1]
            country = place[0]
        if "Date of Publication" in i.values():
            date = list(i.values())[1]
        if "Title" in i.values():
            title = list(i.values())[1]
        if "Publisher" in i.values():
            printer = list(i.values())[1]

    # Nested list with information about each document (One list per document)
    item_data.append([m, city, country, date, title, printer])

# We convert each nested list into a row in the Excel file
row = 1  # Initialisation of the number of row
for data in item_data:
    ws.cell(column=1, row=row, value=data[0])
    ws.cell(column=2, row=row, value=data[1])
    ws.cell(column=3, row=row, value=data[2])
    ws.cell(column=4, row=row, value=data[3])
    ws.cell(column=5, row=row, value=data[4])
    ws.cell(column=6, row=row, value=data[5])
    row += 1  # Incrementation of the row

driver.close()
wb.save(filename=wb_filename)
