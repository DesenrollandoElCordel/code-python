import argparse
import os
import xml.etree.ElementTree as et
from openpyxl import Workbook

# Creation of the argument
parser = argparse.ArgumentParser()
parser.add_argument('--alto', type=str, help="Le chemin du dossier des fichiers xml à charger.")
args = parser.parse_args()

# Creation of the file
wb = Workbook()
wb_filename = ""  # Name of the workbook

# Creation of the worksheet
ws1 = wb.active
ws1.title = "metadata"

# Lists for XML data
list_woodcuts = []

# Iteration through the folders and files
for subdir, dirs, files in os.walk(args.alto):
    for file in files:
        xml_path = os.path.join(subdir, file)  # Path of the XML files
        # print(xml_path)
        label, ext = os.path.splitext(file)  # Splitting of the filename
        # print(label)
        ns = {'alto': 'http://www.loc.gov/standards/alto/ns-v4#'}  # Alto-XML namespace

        if file.endswith('.xml'):
            document = et.parse(xml_path)
            root = document.getroot()
            # print(root)
            woodcutCounter = 0  # Initialisation of the woodcuts number after each iteration

            # Detection of image block (#BT224) in Alto-XML, following the SegmOnto identifier
            for ImgBlock in root.findall('.//alto:TextBlock[@TAGREFS="BT224"]', ns):
                xStart = float(ImgBlock.get('HPOS'))  # Upper Left X (ulx)
                yStart = float(ImgBlock.get('VPOS'))  # Upper Left Y (uly)
                xEnd = float(ImgBlock.get('WIDTH')) + xStart  # Lower right X (lrx) = Width + ulx
                yEnd = float(ImgBlock.get('HEIGHT')) + yStart  # Lower right Y (lry) = Height + uly
                coords = str(xStart) + "," + str(yStart) + "," + str(xEnd) + "," + str(yEnd)  # List of coordinates

                # For each XML file with an image, we get filename information
                for ImgName in root.findall('.//alto:fileName', ns):
                    pliegoName = ImgName.text[:-6]  # Document identifier
                    pageName = ImgName.text[:-4]  # Page Number
                    woodcutCounter = woodcutCounter + 1  # We count the number of images per file
                    woodcutName = 'grabado_v' + ImgName.text[6:-5] + str(woodcutCounter)  # Woodcut identifier

                    # All the information are put in a nested list
                    for i in range(len(ImgBlock)):
                        list_woodcuts.append([woodcutName, pliegoName, pageName, coords])

                    # From the information in list_woodcuts, we create the Excel file
                    row = 1  # Initialisation of the row
                    for x in list_woodcuts:
                        ws1.cell(column=1, row=row, value=x[0])
                        ws1.cell(column=2, row=row, value=x[1])
                        ws1.cell(column=3, row=row, value=x[2])
                        ws1.cell(column=4, row=row, value=x[3])
                        row += 1  # Incrementation of the row

print(list_woodcuts)
# wb.save(filename=wb_filename)
