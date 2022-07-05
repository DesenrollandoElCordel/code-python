import argparse
from openpyxl import load_workbook
import xml.etree.ElementTree as eT
import os
import re

# Creation of the argument
parser = argparse.ArgumentParser()
parser.add_argument('--tei', type=str, help="Path to the XML-TEI folder.")
args = parser.parse_args()

# Loading of the Excel file
wb = load_workbook(filename='../Index_Grabados_Varios.xlsx')

# Selection of the worksheet
ws1 = wb["metadata"]

# Addition of data to XLSX worksheet
for f in os.listdir(args.tei):
    if f.endswith('.xml'):
        xml_path = os.path.join(args.tei, f)  # Path of XML files
        label, ext = os.path.splitext(f)  # Splitting of the filename

        # Namespace declarations
        ns = {'tei': 'http://www.tei-c.org/ns/1.0'}
        eT.register_namespace('', 'http://www.tei-c.org/ns/1.0')

        xml_tree = eT.parse(xml_path)  # Parsing of the XML file
        root = xml_tree.getroot()  # We get the root

        printer = root.find(".//tei:publisher", ns).text  # Name of the printer
        title = root.find(".//tei:fileDesc/tei:titleStmt/tei:title", ns).text.capitalize()  # Title
        date = root.find(".//tei:sourceDesc//tei:publicationStmt/tei:date", ns).text  # Date

        # Iteration through the rows and the second columns (with document name)
        for row in ws1.iter_rows(min_row=2, max_row=127, min_col=2, max_col=2):
            for cell in row:
                number_row = re.split("B", str(cell))  # We split the name of the cell to only get its number
                c = cell.value  # We get the value of each cell
                # If the value of the cell is the same as the name of the XML file,
                # we add information about the title, the date and the printer of the document
                # in the corresponding column and row.
                if c == label:
                    name_cell_columnE = "E" + number_row[1][:-1]  # Concatenation of the name of the column and the number of the row
                    ws1[name_cell_columnE] = title  # We put in the cell the title of the document

                    name_cell_columnF = "F" + number_row[1][:-1]
                    ws1[name_cell_columnF] = date

                    name_cell_columnG = "G" + number_row[1][:-1]
                    ws1[name_cell_columnG] = printer
                else:
                    pass

wb.save(filename='../Index_Grabados_Varios_Update.xlsx')
